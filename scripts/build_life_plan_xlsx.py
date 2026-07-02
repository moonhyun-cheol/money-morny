#!/usr/bin/env python3
"""인생 플랜 Excel(xlsx) 생성."""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from config.life_plan import (  # noqa: E402
    ALLOWANCE_P1,
    ALLOWANCE_P2,
    ALLOWANCE_P1_JULY,
    ALLOWANCE_P2_JULY,
    ANNUAL_INCOME_P1,
    ANNUAL_INCOME_P2,
    BANK_ACCOUNTS,
    REAL_WALLETS,
    ROLE_SLOT_COUNT,
    ACCOUNTS_TO_OPEN,
    CHEONGYAK_BALANCE,
    PHYSICAL_ACCOUNT_COUNT,
    ASSET_ALLOCATION_AUGUST,
    COHAB_P1_TRANSFER,
    COHAB_HUB_ID,
    COHAB_HUB_INSTITUTION,
    CONTRACT_DEPOSIT_TARGET,
    HOUSE_FUND_SOURCES,
    FIXED_COST_ITEMS,
    FIXED_ELECTRIC,
    FIXED_INSURANCE,
    FIXED_JULY_TOTAL,
    FIXED_LOAN_KAKAO,
    FIXED_LOAN_SCHOLAR_INTEREST,
    FIXED_MONTHLY_TOTAL,
    FIXED_NHIS_P2,
    FIXED_P1_ITEMS,
    FIXED_P1_TOTAL,
    FIXED_P2_ITEMS,
    FIXED_P2_TOTAL,
    FIXED_PHONE_KT,
    FIXED_PHONE_P2,
    FIXED_PHONE_SKT,
    FIXED_RENT,
    FIXED_TRANSPORT,
    FIXED_TRANSPORT_P1,
    FIXED_TRANSPORT_P2,
    FOOD_GROCERY,
    FOOD_JULY,
    HOUSING_GOAL,
    JULY_BUFFER,
    JULY_NET_INCOME,
    BELT_COMFORT_FROM,
    BELT_TIGHT_PEAK_UNTIL,
    LIFE_PHASES,
    MARRIAGE_PLAN,
    MONTHLY_INVEST_TOTAL,
    CHILD_SCENARIO_RESULTS,
    CHILD_SCENARIOS,
    CHILD_BIRTH_RECOMMENDED,
    EMERGENCY_FUND_TARGET,
    FIXED_AFTER_MOVE_IN,
    INCOME_GROWTH_NOTE,
    INCOME_GROWTH_RATE,
    MORTGAGE_YEARLY,
    POST_MOVE_COMPARE,
    POST_MOVE_FREE_ALLOCATION,
    POST_MOVE_YEARLY_PROJECTION,
    MORTGAGE_MONTHLY_DD,
    MORTGAGE_PRINCIPAL,
    MOVE_IN_CASH_TARGET,
    FREE_CASH_AFTER_MOVE_IN,
    NET_INCOME_HOUSEHOLD,
    NET_INCOME_P1,
    NET_INCOME_P2,
    NET_INCOME_P1_PROBATION,
    PAYROLL_SCHEDULE_2026,
    P1_PAYDAY_NOTE,
    P1_PROBATION_RATE,
    P1_START_DATE,
    PARKING_P1,
    PARKING_P2,
    P2_FIRST_PAYCHECK,
    P2_TOSS_ACCOUNT_ID,
    PERSON1_NAME,
    PERSON2_NAME,
    PENSION_PLAN_P1,
    PENSION_TAX_CREDIT_RATE,
    SAVE_HOUSE,
    SAVE_HOUSE_P1,
    SAVE_HOUSE_P2,
    SAVE_AFTER_MOVE_IN,
    SAVE_ISA,
    SAVE_PENSION_P1,
    SAVE_VISA_BUFFER_P2,
    SUB_COUPANG,
    SUB_CURSOR_KRW,
    SUB_ICLOUD,
    SUB_LIMBUS,
    SUB_SPOTIFY,
    SUB_YOUTUBE,
    TRANSIT_P1_DAYS,
    TRANSIT_P1_FARE,
    TRANSIT_P2_DAYS,
    TRANSIT_P2_FARE,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    raise

OUT = ROOT / "sheets" / "인생플랜.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _header_row(ws, row: int, values: list) -> None:
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def sheet_overview(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "개요"
    ws["A1"] = "인생 플랜 (운정 59형 · 2인 가구)"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("작성일", date.today().isoformat()),
        ("Person1", f"{PERSON1_NAME} (연 {ANNUAL_INCOME_P1:,}만·{P1_START_DATE.isoformat()} 입사·수습 {int(P1_PROBATION_RATE*100)}%)"),
        ("Person2", f"{PERSON2_NAME} (연 {ANNUAL_INCOME_P2:,}만)"),
        ("혼인", MARRIAGE_PLAN),
        ("주택 목표", HOUSING_GOAL),
        ("허리띠 최대", f"{BELT_TIGHT_PEAK_UNTIL} (이후 완화)"),
        ("숨통 트임", f"{BELT_COMFORT_FROM} · 입주 후 월 여유 ~{FREE_CASH_AFTER_MOVE_IN // 10_000}만"),
        ("월 세후 합산", NET_INCOME_HOUSEHOLD),
        ("월 고정비", "='고정비'!D33"),
        ("월 식비(장보기)", FOOD_GROCERY),
        ("월 용돈", ALLOWANCE_P1 + ALLOWANCE_P2),
        ("월 투자·저축", MONTHLY_INVEST_TOTAL),
        ("계약금 목표", CONTRACT_DEPOSIT_TARGET),
        ("비고", "7월=현철 월급만 · 8월~=자산분배 시트 참고"),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48


def sheet_life_phases(wb: Workbook) -> None:
    ws = wb.create_sheet("라이프페이즈", 2)
    ws["A1"] = "허리띠는 죽을 때까지가 아님 — 집 사기 스프린트 ~5년"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"최대 조임: {BELT_TIGHT_PEAK_UNTIL} · 완화: {BELT_COMFORT_FROM}"
    _header_row(
        ws,
        4,
        ["단계", "이름", "기간", "허리띠", "월 여유(원)", "요약", "할 일"],
    )
    for r, ph in enumerate(LIFE_PHASES, 5):
        ws.cell(row=r, column=1, value=ph["id"])
        ws.cell(row=r, column=2, value=ph["name"])
        ws.cell(row=r, column=3, value=ph["period"])
        ws.cell(row=r, column=4, value=ph["belt_label"])
        free = ph["monthly_free"]
        ws.cell(row=r, column=5, value=free if isinstance(free, int) and free >= 0 else "소득↑ 가정")
        ws.cell(row=r, column=6, value=ph["summary"])
        ws.cell(row=r, column=7, value=ph["tasks"])
    tr = len(LIFE_PHASES) + 6
    ws.cell(row=tr, column=1, value="입주 후 참고")
    ws.cell(
        row=tr,
        column=6,
        value=f"대출 월 {MORTGAGE_MONTHLY_DD:,} · 집마련 저축 중단 · 여유 {FREE_CASH_AFTER_MOVE_IN:,}원/월",
    )
    for col, w in zip("ABCDEFG", [6, 12, 22, 14, 14, 36, 40]):
        ws.column_dimensions[col].width = w


def sheet_fixed_costs(wb: Workbook) -> None:
    """고정비 항목별 — 각 칸 수정 후 월액·합계 자동 계산."""
    ws = wb.create_sheet("고정비", 1)
    ws["A1"] = "고정비 항목별 (노란 칸 = 직접 수정)"
    ws["A1"].font = TITLE_FONT
    _header_row(ws, 3, ["항목", "단가/금액(원)", "수량/일수", "월(원)", "메모"])

    # ── ① 생활 고정비 ──
    ws.cell(row=4, column=1, value="[1] 생활 고정비")
    ws.cell(row=4, column=5, value="월 금액은 D열에 직접 입력")
    direct = [
        (5, "월세+관리비", FIXED_RENT, "60만"),
        (6, "전기", FIXED_ELECTRIC, "추정 5만"),
        (7, "KT", FIXED_PHONE_KT, "FBS"),
        (8, "SKT", FIXED_PHONE_SKT, "지로 6879"),
        (9, "통신(여친)", FIXED_PHONE_P2, ""),
        (10, "보험", FIXED_INSURANCE, "2026-06-04 138,290"),
        (11, "iCloud", SUB_ICLOUD, "카카오페이 등"),
        (12, "쿠팡 와우", SUB_COUPANG, ""),
        (13, "림버스컴퍼니", SUB_LIMBUS, "9900+4900"),
        (14, "Spotify", SUB_SPOTIFY, ""),
        (15, "YouTube Premium", SUB_YOUTUBE, ""),
        (16, "Cursor", SUB_CURSOR_KRW, "Pro+ 12만/월"),
        (17, "장학금 이자", FIXED_LOAN_SCHOLAR_INTEREST, "이자만"),
        (18, "카카오 긴급생활", FIXED_LOAN_KAKAO, "6.7%·36m"),
        (19, "건강보험(여친)", FIXED_NHIS_P2, "직장가입 후"),
    ]
    for r, name, amount, memo in direct:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=4, value=amount)
        ws.cell(row=r, column=5, value=memo)

    ws.cell(row=21, column=1, value="[교통] 정가 (요금×출근일)")
    transit = [
        (22, f"정가-{PERSON1_NAME}", TRANSIT_P1_FARE, TRANSIT_P1_DAYS, "왕복 9800"),
        (23, f"정가-{PERSON2_NAME}", TRANSIT_P2_FARE, TRANSIT_P2_DAYS, "월화수토일·왕복 3500"),
    ]
    for r, name, fare, days, memo in transit:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=fare)
        ws.cell(row=r, column=3, value=days)
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
        ws.cell(row=r, column=5, value=memo)

    ws.cell(row=24, column=1, value="정가 합계")
    ws.cell(row=24, column=4, value="=D22+D23")
    ws.cell(row=25, column=1, value=f"교통-{PERSON1_NAME}")
    ws.cell(row=25, column=4, value=FIXED_TRANSPORT_P1)
    ws.cell(row=25, column=5, value="왕복 9800×22")
    ws.cell(row=26, column=1, value=f"교통-{PERSON2_NAME}")
    ws.cell(row=26, column=4, value=FIXED_TRANSPORT_P2)
    ws.cell(row=26, column=5, value="왕복 3500×22")
    ws.cell(row=27, column=1, value="교통 합계")
    ws.cell(row=27, column=4, value="=D25+D26")
    ws.cell(row=33, column=1, value="고정비 합계")
    ws.cell(row=33, column=4, value="=SUM(D5:D19)+D27")

    for col, w in zip("ABCDE", [20, 14, 12, 14, 40]):
        ws.column_dimensions[col].width = w


def sheet_per_person_fixed(wb: Workbook) -> None:
    ws = wb.create_sheet("인당고정비")
    _header_row(ws, 1, ["담당", "항목", "월(원)", "메모"])
    r = 2
    for name, amount, memo in FIXED_P1_ITEMS:
        ws.cell(row=r, column=1, value=PERSON1_NAME)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=amount)
        ws.cell(row=r, column=4, value=memo)
        r += 1
    ws.cell(row=r, column=2, value="합계")
    ws.cell(row=r, column=3, value=FIXED_P1_TOTAL)
    ws.cell(row=r, column=4, value=f"개인고정 상시 {PARKING_P1:,} · 공동이체 {COHAB_P1_TRANSFER:,}")
    r += 2
    for name, amount, memo in FIXED_P2_ITEMS:
        ws.cell(row=r, column=1, value=PERSON2_NAME)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=amount)
        ws.cell(row=r, column=4, value=memo)
        r += 1
    ws.cell(row=r, column=2, value="합계")
    ws.cell(row=r, column=3, value=FIXED_P2_TOTAL)
    ws.cell(row=r, column=4, value=f"본인고정 {PARKING_P2:,} · 동거분담→{COHAB_HUB_INSTITUTION}")


def sheet_july(wb: Workbook) -> None:
    ws = wb.create_sheet("7월생존")
    ws["A1"] = "2026년 7월 — 수습 세후 ~216만 (6월 근무분·5/18 입사)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"여친 소득 8월 입금(7/1~ 근무) · {P1_PAYDAY_NOTE}"
    _header_row(ws, 4, ["구분", "항목", "월(원)", "메모"])
    rows = [
        ("수입", f"{PERSON1_NAME} 세후 (수습 90%)", JULY_NET_INCOME, "정규 ~240만은 10월~"),
        ("수입", f"{PERSON2_NAME}", 0, f"첫 급여 {P2_FIRST_PAYCHECK}"),
        ("고정비", "가구 (건보 제외·여친 전화·교통 포함)", FIXED_JULY_TOTAL, "소득은 8월~"),
        ("식비", "장보기 (축소)", FOOD_JULY, f"{COHAB_HUB_INSTITUTION}"),
        ("용돈", PERSON1_NAME, ALLOWANCE_P1_JULY, "15→10만"),
        ("용돈", PERSON2_NAME, ALLOWANCE_P2_JULY, "0"),
        ("저축", "ISA·연금", 0, "8월부터 재개"),
        ("잔액", "삼성 CMA", JULY_BUFFER, "집마련·비상"),
    ]
    for r, row in enumerate(rows, 5):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)


def sheet_accounts(wb: Workbook) -> None:
    ws = wb.create_sheet("통장구조")
    ws["A1"] = f"물리 통장 {PHYSICAL_ACCOUNT_COUNT}개 · 역할 슬롯 {ROLE_SLOT_COUNT}개"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"청약 {CHEONGYAK_BALANCE:,}원 · 신규 개설 {len(ACCOUNTS_TO_OPEN)}건"
    _header_row(ws, 4, ["명의", "금융사", "담당 역할"])
    for r, w in enumerate(REAL_WALLETS, 5):
        ws.cell(row=r, column=1, value=w["owner"])
        ws.cell(row=r, column=2, value=w["institution"])
        ws.cell(row=r, column=3, value=w["roles"])
    nr = len(REAL_WALLETS) + 6
    if ACCOUNTS_TO_OPEN:
        ws.cell(row=nr, column=1, value="★ 신규 개설")
        _header_row(ws, nr + 1, ["명의", "금융사", "용도", "메모"])
        for r, row in enumerate(ACCOUNTS_TO_OPEN, nr + 2):
            ws.cell(row=r, column=1, value=row["owner"])
            ws.cell(row=r, column=2, value=row["institution"])
            ws.cell(row=r, column=3, value=row["purpose"])
            ws.cell(row=r, column=4, value=row.get("note", ""))
        sr = nr + len(ACCOUNTS_TO_OPEN) + 3
    else:
        sr = nr
    ws.cell(row=sr, column=1, value="월 이체 상세")
    ws.cell(row=sr, column=3, value="아래 슬롯별 금액")
    _header_row(ws, sr + 1, ["ID", "역할", "명의", "금융사", "월 입금", "용도"])
    for r, acct in enumerate(BANK_ACCOUNTS, sr + 2):
        ws.cell(row=r, column=1, value=acct["id"])
        ws.cell(row=r, column=2, value=acct["role"])
        ws.cell(row=r, column=3, value=acct["owner"])
        ws.cell(row=r, column=4, value=acct["institution"])
        ws.cell(row=r, column=5, value=acct["monthly_in"])
        ws.cell(row=r, column=6, value=acct["purpose"])


def sheet_house_fund(wb: Workbook) -> None:
    ws = wb.create_sheet("집자금인출")
    _header_row(ws, 1, ["상품", "집값 사용", "조건", "비고"])
    for r, row in enumerate(HOUSE_FUND_SOURCES, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)


def sheet_payroll_2026(wb: Workbook) -> None:
    ws = wb.create_sheet("2026급여배분")
    ws["A1"] = f"2026 급여 입금·자산 배분 ({P1_START_DATE.isoformat()} 입사 · 수습 {int(P1_PROBATION_RATE*100)}%)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = P1_PAYDAY_NOTE
    _header_row(
        ws,
        4,
        [
            "입금월",
            "근무월",
            "현철",
            "여친",
            "합계",
            "저축",
            "집마련",
            "비고",
        ],
    )
    for i, row in enumerate(PAYROLL_SCHEDULE_2026, 5):
        ws.cell(row=i, column=1, value=row["label"])
        ws.cell(row=i, column=2, value=f"{row['work_year']}-{int(row['work_month']):02d}")
        ws.cell(row=i, column=3, value=row["p1_net"])
        ws.cell(row=i, column=4, value=row["p2_net"])
        ws.cell(row=i, column=5, value=row["household_net"])
        ws.cell(row=i, column=6, value=row["save_total"])
        ws.cell(row=i, column=7, value=row["save_house"])
        ws.cell(row=i, column=8, value=row["note"])
    for col, w in zip("ABCDEFGH", [10, 10, 12, 12, 12, 12, 12, 28]):
        ws.column_dimensions[col].width = w


def sheet_asset_allocation(wb: Workbook) -> None:
    ws = wb.create_sheet("10월자산분배")
    ws["A1"] = "2026년 10월~ — 급여일+1 자동이체 (세후 ~470만·정규)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "8~9월은 수습·전환 → 「2026급여배분」 시트 참고"
    _header_row(ws, 4, ["순서", "항목", PERSON1_NAME, PERSON2_NAME, "목적 계좌"])
    for i, (name, p1, p2, acct) in enumerate(ASSET_ALLOCATION_AUGUST, 1):
        r = i + 4
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=p1)
        ws.cell(row=r, column=4, value=p2)
        ws.cell(row=r, column=5, value=acct)
    tr = len(ASSET_ALLOCATION_AUGUST) + 4
    ws.cell(row=tr, column=2, value="합계")
    ws.cell(row=tr, column=3, value=NET_INCOME_P1)
    ws.cell(row=tr, column=4, value=NET_INCOME_P2)
    ws.cell(row=tr + 1, column=2, value="가구 합계")
    ws.cell(row=tr + 1, column=3, value=NET_INCOME_HOUSEHOLD)
    ws.cell(row=tr + 2, column=2, value="월 저축 합계")
    ws.cell(row=tr + 2, column=3, value=MONTHLY_INVEST_TOTAL)


def sheet_monthly_budget(wb: Workbook) -> None:
    ws = wb.create_sheet("월예산")
    _header_row(ws, 1, ["구분", "항목", "월(원)", "담당", "메모"])
    data = [
        ("고정비", "월세+관리비", FIXED_RENT, "공동", "→ 고정비 시트"),
        ("고정비", "전기", FIXED_ELECTRIC, "공동", ""),
        ("고정비", "KT+SKT", FIXED_PHONE_KT + FIXED_PHONE_SKT, PERSON1_NAME, ""),
        ("고정비", "통신(여친)", FIXED_PHONE_P2, PERSON2_NAME, ""),
        ("고정비", f"교통-{PERSON1_NAME}", FIXED_TRANSPORT_P1, PERSON1_NAME, "K-패스"),
        ("고정비", f"교통-{PERSON2_NAME}", FIXED_TRANSPORT_P2, PERSON2_NAME, "K-패스"),
        ("고정비", "보험", FIXED_INSURANCE, PERSON1_NAME, "138,290"),
        ("고정비", "구독", SUB_ICLOUD + SUB_COUPANG + SUB_LIMBUS + SUB_SPOTIFY + SUB_YOUTUBE + SUB_CURSOR_KRW, PERSON1_NAME, ""),
        ("고정비", "장학금 이자", FIXED_LOAN_SCHOLAR_INTEREST, PERSON1_NAME, ""),
        ("고정비", "카카오 긴급", FIXED_LOAN_KAKAO, PERSON1_NAME, ""),
        ("고정비", "건강보험(여친)", FIXED_NHIS_P2, PERSON2_NAME, ""),
        ("식비", "주간 장보기", FOOD_GROCERY, COHAB_HUB_INSTITUTION, "공동계좌에서 결제"),
        ("용돈", "개인 용돈", ALLOWANCE_P1, PERSON1_NAME, "취미·의류 등"),
        ("용돈", "개인 용돈", ALLOWANCE_P2, PERSON2_NAME, ""),
        ("저축", "ISA 서민형", SAVE_ISA, PERSON1_NAME, "예금·MMF"),
        ("저축", "집마련 CMA·토스파킹", SAVE_HOUSE, "각자", f"CMA {SAVE_HOUSE_P1:,} + {P2_TOSS_ACCOUNT_ID} {SAVE_HOUSE_P2 + SAVE_VISA_BUFFER_P2:,}"),
        ("저축", "연금저축", SAVE_PENSION_P1, PERSON1_NAME, "세액공제용"),
        ("저축", "비자 비상", SAVE_VISA_BUFFER_P2, PERSON2_NAME, f"{P2_TOSS_ACCOUNT_ID} (집마련과 동일)"),
    ]
    for r, row in enumerate(data, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    total_row = len(data) + 2
    ws.cell(row=total_row, column=2, value="합계")
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
    ws.cell(row=total_row + 1, column=2, value="세후 수입")
    ws.cell(row=total_row + 1, column=3, value=NET_INCOME_HOUSEHOLD)
    ws.cell(row=total_row + 2, column=2, value="잔액")
    ws.cell(row=total_row + 2, column=3, value=f"=C{total_row + 1}-C{total_row}")


def sheet_investments(wb: Workbook) -> None:
    ws = wb.create_sheet("투자종목")
    _header_row(ws, 1, ["계좌유형", "계좌명", "담당", "종목/상품", "월납입", "연한도", "목적", "비고"])
    products = [
        ("ISA", "현철-ISA", PERSON1_NAME, "MMF/예금", SAVE_ISA, 20_000_000, "2029 계약금", "서민형 3년"),
        ("주택청약", "현철-청약", PERSON1_NAME, "종합저축", 20_000, None, "청약 순위", "640만 유지"),
        ("연금저축", "현철-연금", PERSON1_NAME, "연금저축", SAVE_PENSION_P1, 6_000_000, "세액공제", "비혼 개인"),
        ("CMA", "현철-집마련", PERSON1_NAME, "삼성 CMA", SAVE_HOUSE_P1, None, "RP이자", "집값 1순위"),
        ("CMA", "여친-집마련", PERSON2_NAME, "토스 파킹", SAVE_HOUSE_P2, None, "고금리", "집값 1순위"),
        ("현금", "여친-비상", PERSON2_NAME, "비자·비상", SAVE_VISA_BUFFER_P2, None, "D-10/F-6", "TOPIK6"),
    ]
    for r, row in enumerate(products, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)


def sheet_tax(wb: Workbook) -> None:
    ws = wb.create_sheet("세금최적화")
    ws["A1"] = "비혼 기간 · 2인 각각 신고 (합산신고 불가)"
    ws["A1"].font = TITLE_FONT
    _header_row(ws, 3, ["항목", PERSON1_NAME, PERSON2_NAME, "메모"])
    credit = int(PENSION_PLAN_P1 * PENSION_TAX_CREDIT_RATE)
    rows = [
        ("연 소득(만)", ANNUAL_INCOME_P1, ANNUAL_INCOME_P2, "세전"),
        ("연금저축 납입(연)", PENSION_PLAN_P1, 0, "P1 240~400만 권장"),
        ("연금 세액공제(15%)", credit, 0, "P2 환급효과 미미"),
        ("ISA 납입", SAVE_ISA * 12, 0, "세액공제 없음·이자비과세"),
        ("청년도약", "불가", "-", "작년(2025) 근로소득 없음"),
        ("체크카드 소득공제", "생활비·장보기", "알바급여", "전통시장·가맹 30% 등"),
        ("신용카드", "고정비 일부", "-", "15% 구간"),
        ("예상 추가 환급(연)", f"{credit + 50_000:,}원", "~0~10만", "연말정산"),
        ("혼인 후", "-", "-", "신혼특공·디딤돌 8,500만 이하"),
    ]
    for r, row in enumerate(rows, 4):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)


def sheet_timeline(wb: Workbook) -> None:
    ws = wb.create_sheet("타임라인")
    _header_row(ws, 1, ["시기", "구분", "할 일", "목표 잔고(집자금)"])
    milestones = [
        ("2026.07", "P0 생존", f"현철 월급만·고정비 {FIXED_JULY_TOTAL // 10_000}만·ISA/연금 보류", JULY_BUFFER),
        ("2026.08", "P1 저축전성", "급여일 자산분배·여가 0", MONTHLY_INVEST_TOTAL),
        ("2027.02", "비자", "D-10 연장 (TOPIK6)", 300_000),
        ("2028.H2", "P2 혼인", "혼인신고·예식·신혼특공", 40_000_000),
        ("2029.06", "P3 숨통", f"카카오 상환 끝·월 +{FIXED_LOAN_KAKAO:,}", 60_000_000),
        ("2029", "P3 분양", "계약금·중도금(집단대출)", CONTRACT_DEPOSIT_TARGET),
        ("2031~32", "P4 입주", f"디딤돌 {MORTGAGE_PRINCIPAL // 10_000:,}만·여유 ~{FREE_CASH_AFTER_MOVE_IN // 10_000}만/월", MOVE_IN_CASH_TARGET),
        ("2033~", "P5 안정", "용돈·외식 복구·허리띠 풀기", -1),
    ]
    monthly_save = SAVE_HOUSE + SAVE_ISA
    for r, (when, cat, task, _) in enumerate(milestones, 2):
        ws.cell(row=r, column=1, value=when)
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=task)
    ws2 = wb.create_sheet("월별저축")
    _header_row(ws2, 1, ["월차", "년월", "월저축", "누적(집자금)"])
    start = date(2026, 7, 1)
    cum = 0
    for m in range(36):
        y = start.year + (start.month - 1 + m) // 12
        mo = (start.month - 1 + m) % 12 + 1
        add = monthly_save if m > 0 else 0
        cum += add
        ws2.cell(row=m + 2, column=1, value=m)
        ws2.cell(row=m + 2, column=2, value=f"{y}-{mo:02d}")
        ws2.cell(row=m + 2, column=3, value=add)
        ws2.cell(row=m + 2, column=4, value=cum)


def sheet_post_move_savings(wb: Workbook) -> None:
    ws = wb.create_sheet("입주후저축")
    ws["A1"] = "입주 후에도 저축은 계속 — 집마련 슬롯만 여유·아이로 전환"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"소득 가정: {INCOME_GROWTH_NOTE}"

    ws["A4"] = "[1] 지금 vs 입주 직후"
    ws["A4"].font = Font(bold=True)
    _header_row(ws, 5, ["항목", "지금(P1)", "입주 후(P4)", "메모"])
    for i, (name, now, after, memo) in enumerate(POST_MOVE_COMPARE, 6):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=now)
        ws.cell(row=i, column=3, value=after)
        ws.cell(row=i, column=4, value=memo)

    br = 6 + len(POST_MOVE_COMPARE) + 2
    ws.cell(row=br, column=1, value="[2] 여유금 배분 권장 (2031~32)")
    ws.cell(row=br, column=1).font = Font(bold=True)
    _header_row(ws, br + 1, ["용도", "월(원)", "메모"])
    for i, (name, amt, memo) in enumerate(POST_MOVE_FREE_ALLOCATION, br + 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=amt)
        ws.cell(row=i, column=3, value=memo)

    mr = br + 2 + len(POST_MOVE_FREE_ALLOCATION) + 2
    ws.cell(row=mr, column=1, value="[3] 대출 — 이자는 소멸·원금은 자산(집 지분)")
    ws.cell(row=mr, column=1).font = Font(bold=True)
    _header_row(ws, mr + 1, ["연차", "연 납입", "이자(소멸)", "원금(자산)", "잔액"])
    for i, row in enumerate(MORTGAGE_YEARLY, mr + 2):
        ws.cell(row=i, column=1, value=row["year"])
        ws.cell(row=i, column=2, value=row["payment_annual"])
        ws.cell(row=i, column=3, value=row["interest"])
        ws.cell(row=i, column=4, value=row["principal"])
        ws.cell(row=i, column=5, value=row["balance_end"])

    for col, w in zip("ABCDE", [22, 14, 14, 14, 16]):
        ws.column_dimensions[col].width = w


def sheet_child_scenarios(wb: Workbook) -> None:
    ws = wb.create_sheet("출산시나리오")
    ws["A1"] = f"출산 권장 시기: {CHILD_BIRTH_RECOMMENDED} (입주 후 1~2년) · 소득 연 {INCOME_GROWTH_RATE:.0%} 성장"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"비상금 목표 {EMERGENCY_FUND_TARGET:,}원 · 출산 전 아이통장 월 30만 권장"

    _header_row(
        ws,
        4,
        [
            "ID",
            "시나리오",
            "시기",
            "세후수입",
            "고정비",
            "식비",
            "용돈",
            "ISA",
            "연금+비자",
            "육아비",
            "여유",
            "저축+여유",
            "판정",
            "메모",
        ],
    )
    for r, sc in enumerate(CHILD_SCENARIO_RESULTS, 5):
        free = int(sc["free_cash"])
        wealth = int(sc["monthly_wealth"])
        if free >= 200_000:
            verdict = "✅ 여유"
        elif free >= 0:
            verdict = "🟡 빡빡"
        else:
            verdict = "🔴 ISA축소 필요"
        ws.cell(row=r, column=1, value=sc["id"])
        ws.cell(row=r, column=2, value=sc["name"])
        ws.cell(row=r, column=3, value=sc["when"])
        ws.cell(row=r, column=4, value=sc["net_income"])
        ws.cell(row=r, column=5, value=sc["fixed"])
        ws.cell(row=r, column=6, value=sc["food"])
        ws.cell(row=r, column=7, value=sc["allowance"])
        ws.cell(row=r, column=8, value=sc["save_isa"])
        ws.cell(row=r, column=9, value=sc["save_pension"] + int(sc["save_visa"]))
        ws.cell(row=r, column=10, value=sc["child"])
        ws.cell(row=r, column=11, value=free)
        ws.cell(row=r, column=12, value=wealth)
        ws.cell(row=r, column=13, value=verdict)
        ws.cell(row=r, column=14, value=sc["note"])

    tr = 5 + len(CHILD_SCENARIO_RESULTS) + 2
    ws.cell(row=tr, column=1, value="[연도별 투영] 2033 출산(표준) 가정 — ISA 3년차부터 40만")
    ws.cell(row=tr, column=1).font = Font(bold=True)
    _header_row(
        ws,
        tr + 1,
        ["연도", "세후", "육아비", "ISA", "여유", "월 저축+여유"],
    )
    for i, row in enumerate(POST_MOVE_YEARLY_PROJECTION, tr + 2):
        ws.cell(row=i, column=1, value=row["label"])
        ws.cell(row=i, column=2, value=row["net_income"])
        ws.cell(row=i, column=3, value=row["child"])
        ws.cell(row=i, column=4, value=row["save_isa"])
        ws.cell(row=i, column=5, value=row["free_cash"])
        ws.cell(row=i, column=6, value=row["monthly_wealth"])

    for col, w in zip("ABCDEFGHIJKLMN", [5, 12, 10, 10, 10, 10, 10, 8, 10, 10, 10, 12, 10, 28]):
        ws.column_dimensions[col].width = w


def sheet_allowance(wb: Workbook) -> None:
    ws = wb.create_sheet("용돈")
    _header_row(ws, 1, ["담당", "월 용돈", "포함", "미포함"])
    ws.cell(row=2, column=1, value=PERSON1_NAME)
    ws.cell(row=2, column=2, value=ALLOWANCE_P1)
    ws.cell(row=2, column=3, value="개인 취미·의류·PC")
    ws.cell(row=2, column=4, value="장보기·고정비·저축")
    ws.cell(row=3, column=1, value=PERSON2_NAME)
    ws.cell(row=3, column=2, value=ALLOWANCE_P2)
    ws.cell(row=3, column=3, value="개인 소비")
    ws.cell(row=3, column=4, value="공동 식비")


def main() -> None:
    wb = Workbook()
    sheet_overview(wb)
    sheet_life_phases(wb)
    sheet_fixed_costs(wb)
    sheet_per_person_fixed(wb)
    sheet_july(wb)
    sheet_payroll_2026(wb)
    sheet_accounts(wb)
    sheet_house_fund(wb)
    sheet_asset_allocation(wb)
    sheet_monthly_budget(wb)
    sheet_investments(wb)
    sheet_tax(wb)
    sheet_timeline(wb)
    sheet_post_move_savings(wb)
    sheet_child_scenarios(wb)
    sheet_allowance(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"생성: {OUT}")


if __name__ == "__main__":
    main()
