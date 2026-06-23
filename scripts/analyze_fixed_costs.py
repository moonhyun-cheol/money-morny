#!/usr/bin/env python3
"""지출 내역에서 같은 금액 반복 → 고정비 후보 추출."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = Path(__file__).resolve().parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))
DATA_DIR = ROOT / "data"
OUT = ROOT / "sheets" / "고정비분석.xlsx"

DATE_KEYS = ("날짜", "일자", "거래일", "이용일", "승인일", "date", "거래일시")
AMOUNT_KEYS = ("금액", "출금", "지출", "이용금액", "거래금액", "결제금액", "amount", "출금액", "지출금액")
MERCHANT_KEYS = ("적요", "사용처", "가맹점", "거래처", "내용", "memo", "상세", "거래내용", "기재내용")

DATE_PATTERNS = (
    "%Y-%m-%d",
    "%Y.%m.%d",
    "%Y/%m/%d",
    "%Y%m%d",
    "%y-%m-%d",
    "%y.%m.%d",
)


@dataclass
class Transaction:
    dt: date
    amount: int
    merchant: str
    source: str


@dataclass
class FixedCandidate:
    amount: int
    count: int
    month_count: int
    merchants: list[tuple[str, int]]
    dates: list[date]
    sources: list[str]


def _norm_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _parse_amount(raw: object) -> int | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return abs(int(round(float(raw))))
    s = str(raw).strip().replace(",", "").replace("원", "")
    if not s or s in ("-", "0"):
        return None
    sign = -1 if s.startswith("-") or s.startswith("(") else 1
    s = re.sub(r"[^\d.]", "", s)
    if not s:
        return None
    return abs(int(round(float(s))))


def _parse_date(raw: object) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if " " in s:
        s = s.split(" ")[0]
    for fmt in DATE_PATTERNS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.match(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def _find_col(headers: list[str], keys: tuple[str, ...]) -> int | None:
    norm = [_norm_header(h) for h in headers]
    for i, h in enumerate(norm):
        for key in keys:
            if key in h or h in key:
                return i
    return None


def _read_csv(path: Path) -> list[Transaction]:
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = path.read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"인코딩 확인 필요: {path}")

    rows = list(csv.reader(text.splitlines()))
    if not rows:
        return []

    header_idx = 0
    for i, row in enumerate(rows[:20]):
        if _find_col(row, DATE_KEYS) is not None and _find_col(row, AMOUNT_KEYS) is not None:
            header_idx = i
            break

    headers = rows[header_idx]
    date_col = _find_col(headers, DATE_KEYS)
    amount_col = _find_col(headers, AMOUNT_KEYS)
    merchant_col = _find_col(headers, MERCHANT_KEYS)
    if date_col is None or amount_col is None:
        raise ValueError(f"날짜/금액 열을 찾지 못함: {path.name} (헤더: {headers[:8]})")

    out: list[Transaction] = []
    for row in rows[header_idx + 1 :]:
        if len(row) <= max(date_col, amount_col):
            continue
        dt = _parse_date(row[date_col])
        amount = _parse_amount(row[amount_col])
        if dt is None or amount is None or amount == 0:
            continue
        merchant = row[merchant_col].strip() if merchant_col is not None and merchant_col < len(row) else ""
        out.append(Transaction(dt=dt, amount=amount, merchant=merchant, source=path.name))
    return out


def _read_xlsx(path: Path) -> list[Transaction]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return []

    header_idx = 0
    for i, row in enumerate(rows[:20]):
        headers = [str(c or "") for c in row]
        if _find_col(headers, DATE_KEYS) is not None and _find_col(headers, AMOUNT_KEYS) is not None:
            header_idx = i
            break

    headers = [str(c or "") for c in rows[header_idx]]
    date_col = _find_col(headers, DATE_KEYS)
    amount_col = _find_col(headers, AMOUNT_KEYS)
    merchant_col = _find_col(headers, MERCHANT_KEYS)
    if date_col is None or amount_col is None:
        raise ValueError(f"날짜/금액 열을 찾지 못함: {path.name}")

    out: list[Transaction] = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        cells = list(row)
        if len(cells) <= max(date_col, amount_col):
            continue
        dt = _parse_date(cells[date_col])
        amount = _parse_amount(cells[amount_col])
        if dt is None or amount is None or amount == 0:
            continue
        merchant = ""
        if merchant_col is not None and merchant_col < len(cells) and cells[merchant_col]:
            merchant = str(cells[merchant_col]).strip()
        out.append(Transaction(dt=dt, amount=amount, merchant=merchant, source=path.name))
    return out


def load_transactions(paths: list[Path], *, pdf_password: str | None = None) -> list[Transaction]:
    txs: list[Transaction] = []
    for path in paths:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            txs.extend(_read_csv(path))
        elif suffix in (".xlsx", ".xlsm"):
            txs.extend(_read_xlsx(path))
        elif suffix == ".pdf":
            from kb_pdf_parser import read_kb_pdf

            txs.extend(read_kb_pdf(path, pdf_password))
        else:
            print(f"건너뜀 (csv/xlsx/pdf): {path.name}")
    return sorted(txs, key=lambda t: (t.dt, t.amount))


def find_fixed_candidates(
    txs: list[Transaction],
    *,
    min_count: int = 2,
    min_months: int = 1,
    min_amount: int = 1_000,
) -> list[FixedCandidate]:
    by_amount: dict[int, list[Transaction]] = defaultdict(list)
    for tx in txs:
        if tx.amount >= min_amount:
            by_amount[tx.amount].append(tx)

    span_months = len({(t.dt.year, t.dt.month) for t in txs}) if txs else 0
    required_months = min_months if span_months >= 2 else 1

    candidates: list[FixedCandidate] = []
    for amount, group in by_amount.items():
        if len(group) < min_count:
            continue
        months = {(t.dt.year, t.dt.month) for t in group}
        if len(months) < required_months:
            continue
        merchants = Counter(t.merchant or "(적요 없음)" for t in group).most_common(5)
        candidates.append(
            FixedCandidate(
                amount=amount,
                count=len(group),
                month_count=len(months),
                merchants=merchants,
                dates=sorted({t.dt for t in group}),
                sources=sorted({t.source for t in group}),
            )
        )
    return sorted(candidates, key=lambda c: (-c.amount, -c.count))


def _guess_label(merchants: list[tuple[str, int]], amount: int) -> str:
    if not merchants:
        return f"반복 {amount:,}원"
    name = merchants[0][0]
    if len(name) > 24:
        name = name[:24] + "…"
    return name


def write_report(
    txs: list[Transaction],
    candidates: list[FixedCandidate],
    out_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws = wb.active
    ws.title = "고정비후보"
    ws.append(["항목(추정)", "금액(원)", "반복횟수", "발생월수", "대표적요", "파일"])
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    total = 0
    for c in candidates:
        label = _guess_label(c.merchants, c.amount)
        top_m = c.merchants[0][0] if c.merchants else ""
        ws.append([label, c.amount, c.count, c.month_count, top_m, ", ".join(c.sources)])
        total += c.amount

    ws.append([])
    ws.append(["월 추정 합계", total, "", "", "같은 금액 합산(중복 항목 주의)", ""])
    ws.append(["분석 거래 건수", len(txs), "", "", "", ""])

    ws2 = wb.create_sheet("반복상세")
    ws2.append(["금액", "날짜", "적요", "파일"])
    for cell in ws2[1]:
        cell.font = bold
    cand_amounts = {c.amount for c in candidates}
    for tx in txs:
        if tx.amount in cand_amounts:
            ws2.append([tx.amount, tx.dt.isoformat(), tx.merchant, tx.source])

    ws3 = wb.create_sheet("원본요약")
    ws3.append(["파일", "건수", "기간"])
    for cell in ws3[1]:
        cell.font = bold
    by_source: dict[str, list[Transaction]] = defaultdict(list)
    for tx in txs:
        by_source[tx.source].append(tx)
    for src, group in sorted(by_source.items()):
        dts = [t.dt for t in group]
        ws3.append([src, len(group), f"{min(dts)} ~ {max(dts)}"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def print_summary(candidates: list[FixedCandidate], txs: list[Transaction]) -> None:
    if not txs:
        print("거래 내역이 없습니다.")
        return
    print(f"\n=== 고정비 후보 (같은 금액 {len(candidates)}건) ===\n")
    print(f"{'금액':>12}  {'횟수':>4}  {'월':>3}  대표적요")
    print("-" * 56)
    total = 0
    for c in candidates:
        top = c.merchants[0][0] if c.merchants else "-"
        if len(top) > 20:
            top = top[:20] + "…"
        print(f"{c.amount:>12,}  {c.count:>4}  {c.month_count:>3}  {top}")
        total += c.amount
    print("-" * 56)
    print(f"{'합계(중복주의)':>12}  {total:>12,}  ← 같은 금액끼리 더한 값")
    print(f"\n총 거래 {len(txs)}건 분석")


def collect_input_paths(args: argparse.Namespace) -> list[Path]:
    if args.files:
        return [Path(p).resolve() for p in args.files]
    if not DATA_DIR.exists():
        DATA_DIR.mkdir(parents=True, exist_ok=True)
    paths = sorted(DATA_DIR.glob("*.csv")) + sorted(DATA_DIR.glob("*.xlsx")) + sorted(DATA_DIR.glob("*.pdf"))
    return [p for p in paths if not p.name.startswith("~")]


def main() -> int:
    parser = argparse.ArgumentParser(description="같은 금액 반복 → 고정비 후보 추출")
    parser.add_argument("files", nargs="*", help="CSV/XLSX/PDF 경로 (없으면 data/ 폴더 전체)")
    parser.add_argument("--password", "-p", help="PDF 암호 (KB = 생년월일 6자리)")
    parser.add_argument("--min-count", type=int, default=2, help="최소 반복 횟수 (기본 2)")
    parser.add_argument("--min-months", type=int, default=2, help="최소 발생 월수 (기본 2, 1개월치면 1)")
    parser.add_argument("--min-amount", type=int, default=1_000, help="최소 금액 (기본 1,000원)")
    parser.add_argument("-o", "--output", type=Path, default=OUT, help="결과 xlsx 경로")
    args = parser.parse_args()

    paths = collect_input_paths(args)
    if not paths:
        print(f"파일 없음: {DATA_DIR} 에 CSV/XLSX/PDF 를 넣거나 경로를 인자로 주세요.")
        print("  예: python scripts/analyze_fixed_costs.py \"경로/KB.pdf\" --password 991216")
        return 1

    txs: list[Transaction] = []
    for path in paths:
        if not path.exists():
            print(f"없음: {path}")
            continue
        try:
            loaded = load_transactions([path], pdf_password=args.password)
            print(f"  OK {path.name}: {len(loaded)}건")
            txs.extend(loaded)
        except Exception as e:
            print(f"  NG {path.name}: {e}")

    if not txs:
        print("읽은 거래가 없습니다.")
        return 1

    candidates = find_fixed_candidates(
        txs,
        min_count=args.min_count,
        min_months=args.min_months,
        min_amount=args.min_amount,
    )
    print_summary(candidates, txs)
    write_report(txs, candidates, args.output.resolve())
    print(f"\n저장: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
