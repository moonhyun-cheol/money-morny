#!/usr/bin/env python3
"""인생플랜 xlsx 시트 내용 터미널 출력."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from config.life_plan import CHILD_SCENARIO_RESULTS, POST_MOVE_COMPARE, POST_MOVE_YEARLY_PROJECTION

XLSX = ROOT / "sheets" / "인생플랜.xlsx"


def _row_line(row: tuple) -> str:
    cells = [str(c) if c is not None else "" for c in row]
    while cells and cells[-1] == "":
        cells.pop()
    return " | ".join(cells) if any(cells) else ""


def print_sheet(name: str, max_row: int = 30) -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        print(f"[없음] {name}")
        wb.close()
        return
    ws = wb[name]
    print(f"\n{'=' * 64}")
    print(f"  📋 {name}")
    print("=" * 64)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > max_row:
            print("  ...")
            break
        line = _row_line(row)
        if line:
            print(line)
    wb.close()


def main() -> None:
    if not XLSX.exists():
        print(f"없음: {XLSX} — 먼저 build_life_plan_xlsx.py 실행")
        return

    print(f"파일: {XLSX}")
    for s in ["개요", "라이프페이즈", "입주후저축", "출산시나리오"]:
        print_sheet(s, 22 if s != "출산시나리오" else 18)

    print(f"\n{'=' * 64}")
    print("  📊 입주 후 비교 (코드 기준)")
    print("=" * 64)
    print(f"{'항목':<14} {'지금':>12} {'입주후':>12}  메모")
    print("-" * 64)
    for name, now, after, memo in POST_MOVE_COMPARE:
        print(f"{name:<14} {now:>12,} {after:>12,}  {memo}")

    print(f"\n{'=' * 64}")
    print("  👶 출산 시나리오")
    print("=" * 64)
    print(f"{'ID':<4} {'이름':<14} {'시기':<10} {'세후':>9} {'여유':>9} {'저축+여유':>10}  판정")
    print("-" * 64)
    for sc in CHILD_SCENARIO_RESULTS:
        free = int(sc["free_cash"])
        wealth = int(sc["monthly_wealth"])
        if free >= 200_000:
            v = "✅"
        elif free >= 0:
            v = "🟡"
        else:
            v = "🔴"
        print(
            f"{sc['id']:<4} {sc['name']:<14} {sc['when']:<10} "
            f"{int(sc['net_income']):>9,} {free:>9,} {wealth:>10,}  {v}"
        )

    print(f"\n{'=' * 64}")
    print("  📈 연도별 투영 (2033 출산 표준)")
    print("=" * 64)
    print(f"{'연도':<8} {'세후':>10} {'육아비':>8} {'ISA':>8} {'여유':>10} {'저축+여유':>10}")
    print("-" * 64)
    for p in POST_MOVE_YEARLY_PROJECTION:
        print(
            f"{p['label']:<8} {int(p['net_income']):>10,} "
            f"{int(p['child']):>8,} {int(p['save_isa']):>8,} "
            f"{int(p['free_cash']):>10,} {int(p['monthly_wealth']):>10,}"
        )


if __name__ == "__main__":
    main()
