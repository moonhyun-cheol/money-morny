#!/usr/bin/env python3
"""종합 인생라인 Excel 생성 — Google Drive 업로드용."""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from config.life_plan import (  # noqa: E402
    CHILD_SCENARIO_RESULTS,
    LIFE_PHASES,
    POST_MOVE_COMPARE,
)
from config.life_timeline import (  # noqa: E402
    QUARTERLY_MILESTONES,
    TIMELINE_END_YEAR,
    TIMELINE_START_YEAR,
    build_yearly_timeline,
    dashboard_summary,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    raise

OUT = ROOT / "sheets" / "종합인생라인.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1A5276")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(bold=True, size=11)
PHASE_FILLS = {
    "🔴": PatternFill("solid", fgColor="FADBD8"),
    "🟠": PatternFill("solid", fgColor="FDEBD0"),
    "🟡": PatternFill("solid", fgColor="FCF3CF"),
    "🟢": PatternFill("solid", fgColor="D5F5E3"),
    "⚪": PatternFill("solid", fgColor="EBF5FB"),
}


def _header_row(ws, row: int, values: list) -> None:
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _fmt_won(n: int) -> str:
    if n >= 100_000_000:
        return f"{n / 100_000_000:.2f}억"
    if n >= 10_000:
        return f"{n // 10_000}만"
    return f"{n:,}"


def sheet_dashboard(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "대시보드"
    ws["A1"] = "종합 인생라인 — 현철·여친 가구"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"생성일 {date.today().isoformat()} · {TIMELINE_START_YEAR}~{TIMELINE_END_YEAR}"
    for i, (k, v) in enumerate(dashboard_summary(), 4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 52

    r = 4 + len(dashboard_summary()) + 2
    ws.cell(row=r, column=1, value="페이즈 요약").font = SUB_FONT
    _header_row(ws, r + 1, ["단계", "이름", "기간", "허리띠"])
    for i, ph in enumerate(LIFE_PHASES, r + 2):
        ws.cell(row=i, column=1, value=ph["id"])
        ws.cell(row=i, column=2, value=ph["name"])
        ws.cell(row=i, column=3, value=ph["period"])
        ws.cell(row=i, column=4, value=ph["belt_label"])


def sheet_life_line(wb: Workbook) -> None:
    ws = wb.create_sheet("인생라인", 1)
    ws["A1"] = f"연도별 인생라인 ({TIMELINE_START_YEAR}~{TIMELINE_END_YEAR})"
    ws["A1"].font = TITLE_FONT
    headers = [
        "연도",
        "나이",
        "페이즈",
        "허리",
        "단계",
        "월 세후",
        "연 세후",
        "월 저축+여유",
        "월 여유",
        "유동자산",
        "집 지분",
        "대출 잔액",
        "순자산",
        "주거",
        "가족",
        "주요 이벤트",
    ]
    _header_row(ws, 3, headers)
    for r, row in enumerate(build_yearly_timeline(), 4):
        belt = str(row["belt"])
        fill = PHASE_FILLS.get(belt)
        ws.cell(row=r, column=1, value=row["year"])
        ws.cell(row=r, column=2, value=f"만 {row['age']}세")
        ws.cell(row=r, column=3, value=row["phase_id"])
        ws.cell(row=r, column=4, value=belt)
        ws.cell(row=r, column=5, value=row["phase_name"])
        ws.cell(row=r, column=6, value=row["monthly_net"])
        ws.cell(row=r, column=7, value=row["annual_net"])
        ws.cell(row=r, column=8, value=row["monthly_save"])
        ws.cell(row=r, column=9, value=row["monthly_free"])
        ws.cell(row=r, column=10, value=row["liquid_assets"])
        ws.cell(row=r, column=11, value=row["home_equity"])
        ws.cell(row=r, column=12, value=row["mortgage"])
        ws.cell(row=r, column=13, value=row["net_worth"])
        ws.cell(row=r, column=14, value=row["housing"])
        ws.cell(row=r, column=15, value=row["family"])
        ws.cell(row=r, column=16, value=row["events"])
        if fill:
            for c in range(1, 17):
                ws.cell(row=r, column=c).fill = fill
    widths = [6, 8, 8, 5, 12, 10, 11, 12, 10, 12, 12, 12, 12, 10, 10, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def sheet_quarterly(wb: Workbook) -> None:
    ws = wb.create_sheet("분기체크")
    ws["A1"] = "분기별 마일스톤 체크리스트"
    ws["A1"].font = TITLE_FONT
    _header_row(ws, 3, ["분기", "영역", "할 일", "메모", "완료"])
    for r, (q, area, task, memo) in enumerate(QUARTERLY_MILESTONES, 4):
        ws.cell(row=r, column=1, value=q)
        ws.cell(row=r, column=2, value=area)
        ws.cell(row=r, column=3, value=task)
        ws.cell(row=r, column=4, value=memo)
        ws.cell(row=r, column=5, value="")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 24


def sheet_finance_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("재무요약")
    ws["A1"] = "입주 전후 · 출산 시나리오 (인생플랜.xlsx와 동기화)"
    ws["A1"].font = TITLE_FONT
    _header_row(ws, 3, ["항목", "지금", "입주 후", "메모"])
    for r, (name, now, after, memo) in enumerate(POST_MOVE_COMPARE, 4):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=now)
        ws.cell(row=r, column=3, value=after)
        ws.cell(row=r, column=4, value=memo)
    br = 4 + len(POST_MOVE_COMPARE) + 2
    ws.cell(row=br, column=1, value="출산 시나리오").font = SUB_FONT
    _header_row(
        ws,
        br + 1,
        ["ID", "시나리오", "시기", "세후", "여유", "저축+여유", "메모"],
    )
    for r, sc in enumerate(CHILD_SCENARIO_RESULTS, br + 2):
        ws.cell(row=r, column=1, value=sc["id"])
        ws.cell(row=r, column=2, value=sc["name"])
        ws.cell(row=r, column=3, value=sc["when"])
        ws.cell(row=r, column=4, value=sc["net_income"])
        ws.cell(row=r, column=5, value=sc["free_cash"])
        ws.cell(row=r, column=6, value=sc["monthly_wealth"])
        ws.cell(row=r, column=7, value=sc["note"])


def sheet_google_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("구글연동")
    ws["A1"] = "Google 스프레드시트에 추가하는 방법"
    ws["A1"].font = TITLE_FONT
    steps = [
        ("방법 A (가장 쉬움)", ""),
        ("1", "https://drive.google.com 접속"),
        ("2", "「새로 만들기」→「파일 업로드」"),
        ("3", "sheets/종합인생라인.xlsx 선택"),
        ("4", "업로드된 파일 우클릭 →「연결 앱」→「Google 스프레드시트」"),
        ("5", "변환 후 폴더 정리·공유 설정"),
        ("", ""),
        ("방법 B (기존 시트에 합치기)", ""),
        ("1", "Google 스프레드시트 새로 만들기"),
        ("2", "파일 → 가져오기 → 업로드 → 종합인생라인.xlsx"),
        ("3", "「각 시트를 가져오기」선택 → 가져오기"),
        ("", ""),
        ("방법 C (자동 · 재무앱 OAuth)", ""),
        ("1", "config/credentials.json 준비 (setup.py와 동일)"),
        ("2", "python scripts/upload_to_google_sheets.py --oauth"),
        ("3", "기존 시트에 추가: --oauth --id=스프레드시트ID"),
        ("", ""),
        ("주의", "엑셀 서식·색은 구글에서 일부 달라질 수 있음"),
        ("갱신", "로컬에서 build 스크립트 재실행 후 다시 업로드"),
        ("관련 파일", "sheets/인생플랜.xlsx — 월예산·고정비 상세"),
    ]
    for i, (a, b) in enumerate(steps, 3):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 56


def main() -> None:
    wb = Workbook()
    sheet_dashboard(wb)
    sheet_life_line(wb)
    sheet_quarterly(wb)
    sheet_finance_summary(wb)
    sheet_google_guide(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"생성: {OUT}")
    print("Google: Drive 업로드 또는 python scripts/upload_to_google_sheets.py")


if __name__ == "__main__":
    main()
