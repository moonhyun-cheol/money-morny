#!/usr/bin/env python3
"""종합인생라인.xlsx → Google 스프레드시트 업로드."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

XLSX = ROOT / "sheets" / "종합인생라인.xlsx"
CONFIG_DIR = ROOT / "config"
BUILD = ROOT / "scripts" / "build_life_timeline_xlsx.py"


def _print_manual() -> None:
    print(
        """
=== Google 스프레드시트에 추가 ===

[가장 쉬움 — Drive 업로드]
1. python scripts/build_life_timeline_xlsx.py
2. https://drive.google.com → 새로 만들기 → 파일 업로드
3. sheets/종합인생라인.xlsx
4. 우클릭 → 연결 앱 → Google 스프레드시트

[기존 재무 시트에 시트만 추가]
1. Google 스프레드시트 열기 (재무관리 deploy로 만든 것)
2. 파일 → 가져오기 → 업로드 → 종합인생라인.xlsx
3. 「각 시트를 가져오기」선택

[자동 업로드 — OAuth]
1. config/credentials.json 준비 (setup.py와 동일)
2. python scripts/upload_to_google_sheets.py --oauth
"""
    )


def _load_xlsx_rows() -> list[tuple[str, list[list]]]:
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    out: list[tuple[str, list[list]]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [
            ["" if c is None else c for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        out.append((name, rows))
    wb.close()
    return out


def upload_oauth(new_file: bool = True, spreadsheet_id: str = "") -> str:
    """기존 deploy OAuth로 업로드. new_file=True면 새 스프레드시트 생성."""
    from deploy.auth import build_services, get_credentials

    if not XLSX.exists():
        raise FileNotFoundError(f"{XLSX} 없음 — build_life_timeline_xlsx.py 먼저 실행")

    creds = get_credentials(CONFIG_DIR)
    sheets_svc, _, _, _ = build_services(creds)
    data = _load_xlsx_rows()

    if new_file or not spreadsheet_id:
        body = {
            "properties": {"title": "종합인생라인"},
            "sheets": [{"properties": {"title": data[0][0]}}],
        }
        created = sheets_svc.spreadsheets().create(body=body).execute()
        spreadsheet_id = created["spreadsheetId"]
        extra = data[1:]
    else:
        extra = data

    for title, rows in extra:
        sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": title}}}]},
        ).execute()

    for title, rows in data:
        if not rows:
            continue
        sheets_svc.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{title}'!A1",
            valueInputOption="USER_ENTERED",
            body={"values": rows},
        ).execute()

    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"


def main() -> None:
    if not XLSX.exists():
        print(f"없음: {XLSX}\n실행: python scripts/build_life_timeline_xlsx.py")
        sys.exit(1)

    use_oauth = "--oauth" in sys.argv
    sheet_id = ""
    for arg in sys.argv[1:]:
        if arg.startswith("--id="):
            sheet_id = arg.split("=", 1)[1]

    if use_oauth or (CONFIG_DIR / "credentials.json").exists():
        try:
            url = upload_oauth(new_file=not sheet_id, spreadsheet_id=sheet_id)
            print(f"업로드 완료:\n{url}")
            return
        except FileNotFoundError as e:
            print(e)
        except Exception as e:
            print(f"OAuth 업로드 실패: {e}")
            _print_manual()
            sys.exit(1)

    _print_manual()


if __name__ == "__main__":
    main()
